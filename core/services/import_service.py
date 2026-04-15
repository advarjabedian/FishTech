"""
Service for importing customer/profile data from Excel/CSV files.
Extracts business logic from views for testability and reuse.
"""
import io
import csv
import logging
from collections import defaultdict
from ..models import Customer, CustomerProfile, Product

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {'customerid', 'customername', 'itemdescription', 'unittype', 'packsize', 'price'}
OPTIONAL_COLUMNS = {'contactname', 'phone', 'email', 'city', 'state', 'itemid'}

IMPORT_HEADERS = ['CustomerID', 'CustomerName', 'ContactName', 'Phone', 'Email',
                  'City', 'State', 'ItemID', 'ItemDescription', 'UnitType', 'PackSize', 'Price']

IMPORT_SAMPLE_ROW = [1001, 'Example Seafood Co', 'John Smith', '555-1234',
                     'john@example.com', 'Los Angeles', 'CA',
                     '', 'Fresh Salmon Fillet', 'LB', 10, 8.50]

IMPORT_COL_WIDTHS = [12, 30, 20, 15, 25, 15, 8, 10, 35, 10, 10, 10]


def normalize_headers(headers):
    """Map header names to column indices via case/space-insensitive normalization."""
    mapping = {}
    for i, h in enumerate(headers):
        normalized = h.strip().lower().replace(' ', '').replace('_', '')
        mapping[normalized] = i
    return mapping


def parse_file(file):
    """Parse uploaded xlsx or csv. Returns (headers, rows) or raises ValueError."""
    filename = file.name.lower()

    if filename.endswith('.csv'):
        content = file.read().decode('utf-8-sig', errors='ignore')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            raise ValueError("File is empty")
        return rows[0], rows[1:]

    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[str(cell.value or '').strip() for cell in row] for row in ws.iter_rows()]
        wb.close()
        if not all_rows:
            raise ValueError("File is empty")
        return all_rows[0], all_rows[1:]

    else:
        raise ValueError("Unsupported file type. Use .xlsx or .csv")


def validate_and_preview(raw_headers, raw_rows, existing_customer_ids):
    """
    Validate parsed rows and build preview data.
    Returns dict with customer_count, profile_count, errors, customers, rows.
    """
    col = normalize_headers(raw_headers)

    missing = REQUIRED_COLUMNS - set(col.keys())
    if missing:
        raise ValueError(
            f'Missing required columns: {", ".join(missing)}. '
            f'Please use the provided template.'
        )

    customers_map = {}
    valid_rows = []
    errors = []

    for i, row in enumerate(raw_rows, start=2):
        if not any(str(v).strip() for v in row):
            continue

        def get(key, default=''):
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return default
            return str(row[idx]).strip()

        customer_id_str = get('customerid')
        customer_name = get('customername')
        description = get('itemdescription')
        unit_type = get('unittype')

        row_errors = []
        if not customer_id_str:
            row_errors.append('Missing CustomerID')
        if not customer_name:
            row_errors.append('Missing CustomerName')
        if not description:
            row_errors.append('Missing ItemDescription')

        try:
            customer_id = int(float(customer_id_str)) if customer_id_str else None
        except ValueError:
            row_errors.append(f'CustomerID must be a number, got: {customer_id_str}')
            customer_id = None

        try:
            pack_size = float(get('packsize', '1') or '1')
        except ValueError:
            pack_size = 1.0

        try:
            price = float(get('price', '0') or '0')
        except ValueError:
            price = 0.0

        try:
            item_id_str = get('itemid', '')
            comp_item_id = int(float(item_id_str)) if item_id_str else None
        except ValueError:
            comp_item_id = None

        if row_errors:
            errors.append({'row': i, 'message': ', '.join(row_errors)})
            continue

        if customer_id not in customers_map:
            customers_map[customer_id] = {
                'customer_id': customer_id,
                'name': customer_name,
                'contact_name': get('contactname'),
                'phone': get('phone'),
                'email': get('email'),
                'city': get('city'),
                'state': get('state'),
                'is_new': customer_id not in existing_customer_ids,
                'item_count': 0,
            }

        customers_map[customer_id]['item_count'] += 1

        valid_rows.append({
            'customer_id': customer_id,
            'customer_name': customer_name,
            'contact_name': get('contactname'),
            'phone': get('phone'),
            'email': get('email'),
            'city': get('city'),
            'state': get('state'),
            'description': description,
            'unit_type': unit_type,
            'pack_size': pack_size,
            'price': price,
            'comp_item_id': comp_item_id,
        })

    customers_list = list(customers_map.values())
    new_count = sum(1 for c in customers_list if c['is_new'])

    return {
        'customer_count': len(customers_list),
        'new_customers': new_count,
        'existing_customers': len(customers_list) - new_count,
        'profile_count': len(valid_rows),
        'error_count': len(errors),
        'customers': customers_list,
        'errors': errors,
        'rows': valid_rows,
    }


def execute_import(tenant, rows):
    """
    Save validated import rows into Customer + CustomerProfile tables.
    Returns dict with customers_created, customers_updated, profiles_created.
    """
    customers_created = 0
    customers_updated = 0
    profiles_created = 0

    by_customer = defaultdict(list)
    for row in rows:
        by_customer[row['customer_id']].append(row)

    for customer_id, items in by_customer.items():
        first = items[0]

        customer, created = Customer.all_objects.update_or_create(
            tenant=tenant,
            customer_id=customer_id,
            defaults={
                'name': first['customer_name'],
                'contact_name': first['contact_name'] or '',
                'phone': first['phone'] or '',
                'email': first['email'] or '',
                'city': first['city'] or '',
                'state': first['state'] or '',
            }
        )

        if created:
            customers_created += 1
        else:
            customers_updated += 1

        CustomerProfile.all_objects.filter(tenant=tenant, customer=customer).delete()

        for item in items:
            try:
                pack_size = float(item['pack_size']) if item.get('pack_size') else None
            except (ValueError, TypeError):
                pack_size = None
            try:
                price = float(item['price']) if item.get('price') else None
            except (ValueError, TypeError):
                price = None

            product, _ = Product.objects.get_or_create(
                tenant=tenant,
                description=item['description'],
                defaults={
                    'unit_type': item['unit_type'] or '',
                    'pack_size': pack_size,
                    'default_price': price,
                }
            )

            CustomerProfile.objects.create(
                tenant=tenant,
                customer=customer,
                product=product,
                description=item['description'],
                unit_type=item['unit_type'] or '',
                pack_size=item['pack_size'],
                sales_price=item['price'],
                comp_item_id=item.get('comp_item_id'),
                is_active=True,
            )
            profiles_created += 1

    logger.info(
        f"Tenant {tenant.name}: Import complete — "
        f"{customers_created} created, {customers_updated} updated, {profiles_created} profiles"
    )

    return {
        'customers_created': customers_created,
        'customers_updated': customers_updated,
        'profiles_created': profiles_created,
    }


def generate_template_workbook():
    """Generate an Excel workbook for the import template. Returns openpyxl Workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import'

    for col, header in enumerate(IMPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')

    ws.append(IMPORT_SAMPLE_ROW)

    for i, w in enumerate(IMPORT_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return wb
