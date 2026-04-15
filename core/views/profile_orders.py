from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from ..models import Customer, CustomerProfile, ProductImage, Product, SalesOrder, SalesOrderItem, get_current_tenant
import json
import logging
from ..models import Tenant, InboundMessage

logger = logging.getLogger(__name__)


@login_required
def customer_list(request):
    customers = Customer.objects.order_by('name')
    return render(request, 'core/Orders/customer_list.html', {'customers': customers})


@login_required
def products_page(request):
    """View tenant product catalog and customer assignments"""
    from django.db.models import Prefetch
    tenant = get_current_tenant()
    products = Product.objects.filter(is_active=True).prefetch_related('images').order_by('sort_order', 'description')
    customers = Customer.objects.prefetch_related(
        Prefetch('profiles', queryset=CustomerProfile.objects.select_related('product').prefetch_related('product__images'))
    ).order_by('name')
    return render(request, 'core/Orders/products.html', {
        'customers': customers,
        'products': products,
    })


def profile_order_form(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    profiles = CustomerProfile.objects.filter(customer=customer, is_active=True).order_by('description')
    return render(request, 'core/Orders/profile_order_form.html', {
        'customer': customer,
        'profiles': profiles,
    })


@login_required
def import_profiles(request):
    """Upload + preview page for importing customer/profile data from Excel"""
    return render(request, 'core/Orders/import_profiles.html')


@login_required
@require_POST
def import_profiles_preview(request):
    """Parse uploaded Excel/CSV and return preview data as JSON"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    filename = uploaded_file.name.lower()

    try:
        if filename.endswith('.csv'):
            import csv, io
            content = uploaded_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
        else:
            return JsonResponse({'error': 'File must be .csv, .xlsx, or .xls'}, status=400)

        if not rows:
            return JsonResponse({'error': 'File is empty'}, status=400)

        # Normalize column names to handle variations
        COL_MAP = {
            'customerid': 'CustomerID', 'customer_id': 'CustomerID', 'custid': 'CustomerID',
            'customername': 'CustomerName', 'customer_name': 'CustomerName', 'customer': 'CustomerName',
            'contactname': 'ContactName', 'contact_name': 'ContactName', 'contact': 'ContactName',
            'customercontactname': 'ContactName',
            'phone': 'Phone', 'customercell': 'Phone', 'cell': 'Phone',
            'email': 'Email', 'customeremail': 'Email', 'emailaddress': 'Email',
            'city': 'City',
            'state': 'State',
            'itemid': 'ItemID', 'item_id': 'ItemID', 'compitemid': 'ItemID',
            'itemdescription': 'ItemDescription', 'item_description': 'ItemDescription',
            'description': 'ItemDescription', 'profiledescription': 'ItemDescription', 'item': 'ItemDescription',
            'unittype': 'UnitType', 'unit_type': 'UnitType', 'unit': 'UnitType', 'unitytype': 'UnitType',
            'packsize': 'PackSize', 'pack_size': 'PackSize', 'pack': 'PackSize', 'packsizepd': 'PackSize',
            'price': 'Price', 'salesprice': 'Price', 'salesprice1': 'Price', 'sales_price': 'Price',
        }

        def normalize_row(row):
            return {COL_MAP.get(k.lower().replace(' ', '').replace('-', ''), k): v for k, v in row.items()}

        rows = [normalize_row(r) for r in rows]

        # Check required columns
        required = {'CustomerID', 'CustomerName', 'ItemDescription'}
        missing = required - set(rows[0].keys())
        if missing:
            return JsonResponse({
                'error': f'Missing required columns: {", ".join(missing)}. Columns found: {", ".join(rows[0].keys())}'
            }, status=400)

        customers_seen = {}
        preview_rows = []
        errors = []

        for i, row in enumerate(rows, start=2):
            cust_id = str(row.get('CustomerID', '')).strip()
            cust_name = str(row.get('CustomerName', '')).strip()
            item_desc = str(row.get('ItemDescription', '')).strip()

            if not cust_id or not cust_name:
                errors.append(f'Row {i}: missing CustomerID or CustomerName')
                continue
            if not item_desc:
                errors.append(f'Row {i}: missing ItemDescription')
                continue

            if cust_id not in customers_seen:
                customers_seen[cust_id] = cust_name

            preview_rows.append({
                'customer_id': cust_id,
                'customer_name': cust_name,
                'contact_name': row.get('ContactName', ''),
                'phone': row.get('Phone', ''),
                'email': row.get('Email', ''),
                'city': row.get('City', ''),
                'state': row.get('State', ''),
                'item_id': row.get('ItemID', ''),
                'item_description': item_desc,
                'unit_type': row.get('UnitType', ''),
                'pack_size': row.get('PackSize', ''),
                'price': row.get('Price', ''),
            })

        return JsonResponse({
            'success': True,
            'customer_count': len(customers_seen),
            'item_count': len(preview_rows),
            'errors': errors[:20],
            'rows': preview_rows[:200],
            'total_rows': len(preview_rows),
        })

    except Exception as e:
        logger.error(f"Import preview error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def import_profiles_confirm(request):
    """Save parsed rows into Customer + CustomerProfile tables"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)

    try:
        data = json.loads(request.body)
        rows = data.get('rows', [])

        if not rows:
            return JsonResponse({'error': 'No data to import'}, status=400)

        from django.db.models import Max
        max_cust = Customer.all_objects.filter(tenant=tenant).aggregate(Max('customer_id'))['customer_id__max']
        next_cust_id = (max_cust + 1) if max_cust else 1

        customers_created = 0
        customers_updated = 0
        profiles_created = 0
        profiles_skipped = 0
        cust_id_to_obj = {}

        for row in rows:
            raw_cust_id = str(row.get('customer_id', '')).strip()
            cust_name = str(row.get('customer_name', '')).strip()
            if not raw_cust_id or not cust_name:
                continue

            if raw_cust_id not in cust_id_to_obj:
                try:
                    customer = Customer.objects.get(name=cust_name)
                    updated = False
                    for field, key in [('phone', 'phone'), ('email', 'email'), ('city', 'city'),
                                       ('state', 'state'), ('contact_name', 'contact_name')]:
                        if row.get(key) and not getattr(customer, field):
                            setattr(customer, field, row[key])
                            updated = True
                    if updated:
                        customer.save()
                        customers_updated += 1
                except Customer.DoesNotExist:
                    try:
                        ext_id = int(raw_cust_id)
                    except ValueError:
                        ext_id = next_cust_id
                        next_cust_id += 1

                    customer = Customer.objects.create(
                        tenant=tenant,
                        customer_id=ext_id,
                        name=cust_name,
                        contact_name=row.get('contact_name', ''),
                        phone=row.get('phone', ''),
                        email=row.get('email', ''),
                        city=row.get('city', ''),
                        state=row.get('state', ''),
                    )
                    customers_created += 1

                cust_id_to_obj[raw_cust_id] = customer

            customer = cust_id_to_obj[raw_cust_id]
            item_desc = str(row.get('item_description', '')).strip()
            if not item_desc:
                continue

            if CustomerProfile.objects.filter(customer=customer, description=item_desc).exists():
                profiles_skipped += 1
                continue

            try:
                pack_size = float(row.get('pack_size') or 1)
            except (ValueError, TypeError):
                pack_size = None

            try:
                price = float(str(row.get('price') or 0).replace('$', '').replace(',', ''))
            except (ValueError, TypeError):
                price = None

            try:
                comp_item_id = int(row.get('item_id') or 0) or None
            except (ValueError, TypeError):
                comp_item_id = None

            # Get or create product catalog entry
            product, _ = Product.objects.get_or_create(
                tenant=tenant,
                description=item_desc,
                defaults={
                    'unit_type': row.get('unit_type', ''),
                    'pack_size': pack_size,
                    'default_price': price,
                }
            )

            CustomerProfile.objects.create(
                tenant=tenant,
                customer=customer,
                product=product,
                description=item_desc,
                unit_type=row.get('unit_type', ''),
                pack_size=pack_size,
                sales_price=price,
                comp_item_id=comp_item_id,
                is_active=True,
            )
            profiles_created += 1

        return JsonResponse({
            'success': True,
            'message': (
                f'Import complete: {customers_created} customers created, '
                f'{customers_updated} updated, '
                f'{profiles_created} profile items added, '
                f'{profiles_skipped} duplicates skipped.'
            )
        })

    except Exception as e:
        logger.error(f"Import confirm error: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def download_import_template(request):
    """Download a pre-formatted Excel import template"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Template'

    headers = ['CustomerID', 'CustomerName', 'ContactName', 'Phone', 'Email',
               'City', 'State', 'ItemID', 'ItemDescription', 'UnitType', 'PackSize', 'Price']
    notes = [
        'Required. Your customer number.',
        'Required. Full customer name.',
        'Optional.',
        'Optional.',
        'Optional.',
        'Optional.',
        'Optional. 2-letter state.',
        'Optional. Your product ID.',
        'Required. Product description.',
        'Optional. e.g. LB, CS, EA',
        'Optional. e.g. 10, 25.5',
        'Optional. e.g. 4.99',
    ]
    sample = ['1001', 'ABC Seafood Co', 'John Smith', '555-1234', 'john@abc.com',
              'Los Angeles', 'CA', '501', 'Atlantic Salmon 10lb', 'CS', '10', '45.00']

    header_fill = PatternFill(start_color='4a90e2', end_color='4a90e2', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    note_font = Font(color='888888', italic=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 20

    for col, note in enumerate(notes, 1):
        ws.cell(row=2, column=col, value=note).font = note_font

    for col, val in enumerate(sample, 1):
        ws.cell(row=3, column=col, value=val)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="fishteck_import_template.xlsx"'
    wb.save(response)
    return response

@require_POST
def submit_profile_order(request):
    try:
        data = json.loads(request.body)
        order_data = data.get('orderData', {})
        order_items = data.get('orderItems', [])

        if not order_items:
            return JsonResponse({'error': 'No items in order'}, status=400)

        customer = get_object_or_404(Customer, id=order_data.get('customerId'))
        tenant = customer.tenant

        # Generate next order number
        last_so = SalesOrder.all_objects.filter(tenant=tenant).order_by('-id').first()
        next_num = 10001
        if last_so:
            try:
                next_num = int(last_so.order_number) + 1
            except (ValueError, TypeError):
                next_num = last_so.id + 10001

        import datetime
        from django.utils.dateparse import parse_date
        dispatch_date_str = order_data.get('dispatchDate', '')
        try:
            dispatch_date = datetime.datetime.strptime(dispatch_date_str, '%m/%d/%Y').date()
        except ValueError:
            dispatch_date = parse_date(dispatch_date_str)

        so = SalesOrder.objects.create(
            tenant=tenant,
            order_number=str(next_num),
            customer=customer,
            customer_name=customer.name,
            order_date=dispatch_date,
            delivery_date=dispatch_date,
            notes=order_data.get('comments', ''),
            po_number=order_data.get('customerPO', ''),
            order_status='open',
        )

        for item in order_items:
            qty = item.get('quantity', 0)
            price = item.get('price', 0)
            pack = item.get('packSize', 1)
            SalesOrderItem.objects.create(
                tenant=tenant,
                sales_order=so,
                description=item.get('name', ''),
                quantity=qty,
                unit_price=price,
                amount=float(qty or 0) * float(pack or 1) * float(price or 0),
                notes=item.get('instructions', ''),
            )

        return JsonResponse({'success': True, 'order_id': so.order_number,
                             'message': f'Order SO-{so.order_number} submitted successfully'})

    except Exception as e:
        logger.error(f"Error submitting profile order: {e}")
        return JsonResponse({'error': str(e)}, status=500)

# =============================================================================
# IMPORT VIEWS
# =============================================================================

import io
import csv
from django.http import HttpResponse


REQUIRED_COLUMNS = {'customerid', 'customername', 'itemdescription', 'unittype', 'packsize', 'price'}
OPTIONAL_COLUMNS = {'contactname', 'phone', 'email', 'city', 'state', 'itemid'}


def _normalize_headers(headers):
    """Map header names to normalized keys"""
    mapping = {}
    for i, h in enumerate(headers):
        normalized = h.strip().lower().replace(' ', '').replace('_', '')
        mapping[normalized] = i
    return mapping


def _parse_file(file):
    """Parse uploaded xlsx or csv, return (headers, rows) or raise ValueError"""
    filename = file.name.lower()

    if filename.endswith('.csv'):
        content = file.read().decode('utf-8-sig', errors='ignore')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            raise ValueError("File is empty")
        return rows[0], rows[1:]

    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[str(cell.value or '').strip() for cell in row] for row in ws.iter_rows()]
        wb.close()
        if not all_rows:
            raise ValueError("File is empty")
        return all_rows[0], all_rows[1:]

    else:
        raise ValueError("Unsupported file type. Use .xlsx or .csv")


@login_required
def import_customers_page(request):
    return render(request, 'core/Orders/import_customers.html')


@login_required
def download_import_template(request):
    """Return a blank Excel template for customer/profile import"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Import'

        headers = ['CustomerID', 'CustomerName', 'ContactName', 'Phone', 'Email',
                   'City', 'State', 'ItemID', 'ItemDescription', 'UnitType', 'PackSize', 'Price']

        # Style header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')

        # Example row
        ws.append([1001, 'Example Seafood Co', 'John Smith', '555-1234',
                   'john@example.com', 'Los Angeles', 'CA',
                   '', 'Fresh Salmon Fillet', 'LB', 10, 8.50])

        # Column widths
        widths = [12, 30, 20, 15, 25, 15, 8, 10, 35, 10, 10, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="fishteck_import_template.xlsx"'
        return response

    except ImportError:
        # Fallback to CSV if openpyxl not available
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fishteck_import_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['CustomerID', 'CustomerName', 'ContactName', 'Phone', 'Email',
                         'City', 'State', 'ItemID', 'ItemDescription', 'UnitType', 'PackSize', 'Price'])
        writer.writerow([1001, 'Example Seafood Co', 'John Smith', '555-1234',
                         'john@example.com', 'Los Angeles', 'CA',
                         '', 'Fresh Salmon Fillet', 'LB', 10, 8.50])
        return response


@login_required
@require_POST
def import_preview(request):
    """Parse uploaded file and return preview data without saving"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    try:
        raw_headers, raw_rows = _parse_file(file)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)

    col = _normalize_headers(raw_headers)

    # Check required columns
    missing = REQUIRED_COLUMNS - set(col.keys())
    if missing:
        return JsonResponse({
            'error': f'Missing required columns: {", ".join(missing)}. '
                     f'Please use the provided template.'
        }, status=400)

    # Get existing customer IDs for this tenant
    existing_ids = set(
        Customer.all_objects.filter(tenant=tenant)
        .values_list('customer_id', flat=True)
    )

    customers_map = {}  # customer_id -> customer dict
    valid_rows = []
    errors = []

    for i, row in enumerate(raw_rows, start=2):
        # Skip empty rows
        if not any(str(v).strip() for v in row):
            continue

        def get(key, default=''):
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return default
            return str(row[idx]).strip()

        customer_id_str = get('customerid')
        customer_name = get('customername')
        description = get('itemdescription')
        unit_type = get('unittype')

        # Validate required fields
        row_errors = []
        if not customer_id_str:
            row_errors.append('Missing CustomerID')
        if not customer_name:
            row_errors.append('Missing CustomerName')
        if not description:
            row_errors.append('Missing ItemDescription')

        try:
            customer_id = int(float(customer_id_str)) if customer_id_str else None
        except ValueError:
            row_errors.append(f'CustomerID must be a number, got: {customer_id_str}')
            customer_id = None

        try:
            pack_size = float(get('packsize', '1') or '1')
        except ValueError:
            pack_size = 1.0

        try:
            price = float(get('price', '0') or '0')
        except ValueError:
            price = 0.0

        try:
            item_id_str = get('itemid', '')
            comp_item_id = int(float(item_id_str)) if item_id_str else None
        except ValueError:
            comp_item_id = None

        if row_errors:
            errors.append({'row': i, 'message': ', '.join(row_errors)})
            continue

        # Build customer record
        if customer_id not in customers_map:
            customers_map[customer_id] = {
                'customer_id': customer_id,
                'name': customer_name,
                'contact_name': get('contactname'),
                'phone': get('phone'),
                'email': get('email'),
                'city': get('city'),
                'state': get('state'),
                'is_new': customer_id not in existing_ids,
                'item_count': 0,
            }

        customers_map[customer_id]['item_count'] += 1

        valid_rows.append({
            'customer_id': customer_id,
            'customer_name': customer_name,
            'contact_name': get('contactname'),
            'phone': get('phone'),
            'email': get('email'),
            'city': get('city'),
            'state': get('state'),
            'description': description,
            'unit_type': unit_type,
            'pack_size': pack_size,
            'price': price,
            'comp_item_id': comp_item_id,
        })

    customers_list = list(customers_map.values())
    new_count = sum(1 for c in customers_list if c['is_new'])

    return JsonResponse({
        'customer_count': len(customers_list),
        'new_customers': new_count,
        'existing_customers': len(customers_list) - new_count,
        'profile_count': len(valid_rows),
        'error_count': len(errors),
        'customers': customers_list,
        'errors': errors,
        'rows': valid_rows,  # passed back to confirm step
    })


@login_required
@require_POST
def import_confirm(request):
    """Save previewed import data"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)

    try:
        data = json.loads(request.body)
        rows = data.get('rows', [])

        if not rows:
            return JsonResponse({'error': 'No rows to import'}, status=400)

        customers_created = 0
        customers_updated = 0
        profiles_created = 0

        # Group rows by customer_id
        from collections import defaultdict
        by_customer = defaultdict(list)
        for row in rows:
            by_customer[row['customer_id']].append(row)

        for customer_id, items in by_customer.items():
            first = items[0]

            # Create or update customer
            customer, created = Customer.all_objects.update_or_create(
                tenant=tenant,
                customer_id=customer_id,
                defaults={
                    'name': first['customer_name'],
                    'contact_name': first['contact_name'] or '',
                    'phone': first['phone'] or '',
                    'email': first['email'] or '',
                    'city': first['city'] or '',
                    'state': first['state'] or '',
                }
            )

            if created:
                customers_created += 1
            else:
                customers_updated += 1

            # Delete existing profiles for this customer before re-importing
            CustomerProfile.all_objects.filter(tenant=tenant, customer=customer).delete()

            # Create profile items
            for item in items:
                try:
                    pack_size = float(item['pack_size']) if item.get('pack_size') else None
                except (ValueError, TypeError):
                    pack_size = None
                try:
                    price = float(item['price']) if item.get('price') else None
                except (ValueError, TypeError):
                    price = None

                # Get or create product catalog entry
                product, _ = Product.objects.get_or_create(
                    tenant=tenant,
                    description=item['description'],
                    defaults={
                        'unit_type': item['unit_type'] or '',
                        'pack_size': pack_size,
                        'default_price': price,
                    }
                )

                CustomerProfile.objects.create(
                    tenant=tenant,
                    customer=customer,
                    product=product,
                    description=item['description'],
                    unit_type=item['unit_type'] or '',
                    pack_size=item['pack_size'],
                    sales_price=item['price'],
                    comp_item_id=item.get('comp_item_id'),
                    is_active=True,
                )
                profiles_created += 1

        logger.info(
            f"Tenant {tenant.name}: Import complete — "
            f"{customers_created} created, {customers_updated} updated, {profiles_created} profiles"
        )

        return JsonResponse({
            'success': True,
            'customers_created': customers_created,
            'customers_updated': customers_updated,
            'profiles_created': profiles_created,
        })

    except Exception as e:
        logger.error(f"Import confirm error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def add_customer_api(request):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)

        from django.db.models import Max
        max_id = Customer.all_objects.filter(tenant=tenant).aggregate(Max('customer_id'))['customer_id__max']
        next_id = (max_id + 1) if max_id else 1

        customer = Customer.objects.create(
            tenant=tenant,
            customer_id=next_id,
            name=name,
            contact_name=data.get('contact_name', ''),
            phone=data.get('phone', ''),
            email=data.get('email', ''),
            address=data.get('address', ''),
            city=data.get('city', ''),
            state=data.get('state', ''),
            zipcode=data.get('zipcode', ''),
        )
        return JsonResponse({'success': True, 'id': customer.id, 'name': customer.name})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def update_customer_api(request, customer_id):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        customer = get_object_or_404(Customer, id=customer_id, tenant=tenant)
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        customer.name = name
        for field in ('contact_name', 'phone', 'email', 'address', 'city', 'state', 'zipcode'):
            if field in data:
                setattr(customer, field, data[field].strip() if isinstance(data[field], str) else data[field])
        customer.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def delete_customer_api(request, customer_id):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        customer = get_object_or_404(Customer, id=customer_id, tenant=tenant)
        customer.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def add_profile_item_api(request, customer_id):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        data = json.loads(request.body)
        description = data.get('description', '').strip()
        if not description:
            return JsonResponse({'error': 'Description is required'}, status=400)

        profile = CustomerProfile.objects.create(
            tenant=tenant,
            customer=customer,
            description=description,
            unit_type=data.get('unit_type', ''),
            pack_size=float(data.get('pack_size') or 1),
            sales_price=float(data.get('sales_price') or 0),
            is_active=True,
        )
        return JsonResponse({
            'success': True,
            'id': profile.id,
            'description': profile.description,
            'unit_type': profile.unit_type,
            'pack_size': str(profile.pack_size),
            'sales_price': str(profile.sales_price),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def update_profile_item_api(request, profile_id):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        profile = get_object_or_404(CustomerProfile, id=profile_id)
        data = json.loads(request.body)
        description = data.get('description', '').strip()
        if not description:
            return JsonResponse({'error': 'Description is required'}, status=400)
        profile.description = description
        profile.unit_type = data.get('unit_type', '')
        profile.pack_size = float(data.get('pack_size') or 1)
        profile.sales_price = float(data.get('sales_price') or 0)
        profile.save()
        return JsonResponse({
            'success': True,
            'id': profile.id,
            'description': profile.description,
            'unit_type': profile.unit_type,
            'pack_size': str(profile.pack_size),
            'sales_price': str(profile.sales_price),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def delete_profile_item_api(request, profile_id):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        profile = get_object_or_404(CustomerProfile, id=profile_id)
        profile.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def upload_product_image(request, product_id):
    """Upload an image (slot 1, 2, or 3) for a product"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        product = get_object_or_404(Product, id=product_id)
        slot = int(request.POST.get('slot', 1))
        if slot not in (1, 2, 3):
            return JsonResponse({'error': 'Slot must be 1, 2, or 3'}, status=400)

        image_file = request.FILES.get('image')
        if not image_file:
            return JsonResponse({'error': 'No image provided'}, status=400)

        # Delete existing image in this slot
        ProductImage.objects.filter(product=product, slot=slot).delete()

        img = ProductImage.objects.create(
            product=product,
            slot=slot,
            image=image_file,
        )
        return JsonResponse({'success': True, 'url': img.image.url, 'slot': slot})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def delete_product_image(request, product_id, slot):
    """Delete a product image by slot"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        product = get_object_or_404(Product, id=product_id)
        img = ProductImage.objects.filter(product=product, slot=slot).first()
        if img:
            img.image.delete(save=False)
            img.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_product_images(request, product_id):
    """Get all images for a tenant product"""
    try:
        product = get_object_or_404(Product, id=product_id)
        images = []
        for img in ProductImage.objects.filter(product=product).order_by('slot'):
            images.append({'slot': img.slot, 'url': img.image.url})
        return JsonResponse({'success': True, 'images': images})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# TENANT PRODUCT CATALOG
# =============================================================================

@login_required
def get_tenant_products_api(request):
    """List all products in the catalog"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    products = Product.objects.order_by('sort_order', 'description')
    data = [{
        'id': p.id,
        'description': p.description,
        'unit_type': p.unit_type,
        'pack_size': str(p.pack_size) if p.pack_size else '',
        'default_price': str(p.default_price) if p.default_price else '',
        'is_active': p.is_active,
    } for p in products]
    return JsonResponse({'success': True, 'products': data})


@login_required
@require_POST
def add_tenant_product_api(request):
    """Add a product to the catalog"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        data = json.loads(request.body)
        description = data.get('description', '').strip()
        if not description:
            return JsonResponse({'error': 'Description is required'}, status=400)
        if Product.objects.filter(description=description).exists():
            return JsonResponse({'error': 'A product with this description already exists'}, status=400)
        product = Product.objects.create(
            tenant=tenant,
            description=description,
            unit_type=data.get('unit_type', ''),
            pack_size=float(data.get('pack_size') or 0) or None,
            default_price=float(data.get('default_price') or 0) or None,
            is_active=True,
        )
        return JsonResponse({
            'success': True,
            'id': product.id,
            'description': product.description,
            'unit_type': product.unit_type,
            'pack_size': str(product.pack_size) if product.pack_size else '',
            'default_price': str(product.default_price) if product.default_price else '',
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def update_tenant_product_api(request, product_id):
    """Update a product"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        product = get_object_or_404(Product, id=product_id)
        data = json.loads(request.body)
        description = data.get('description', '').strip()
        if not description:
            return JsonResponse({'error': 'Description is required'}, status=400)
        dup = Product.objects.filter(description=description).exclude(id=product_id).exists()
        if dup:
            return JsonResponse({'error': 'A product with this description already exists'}, status=400)
        product.description = description
        product.unit_type = data.get('unit_type', '')
        product.pack_size = float(data.get('pack_size') or 0) or None
        product.default_price = float(data.get('default_price') or 0) or None
        product.save()
        return JsonResponse({
            'success': True,
            'id': product.id,
            'description': product.description,
            'unit_type': product.unit_type,
            'pack_size': str(product.pack_size) if product.pack_size else '',
            'default_price': str(product.default_price) if product.default_price else '',
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def delete_tenant_product_api(request, product_id):
    """Delete a product (and unlink from customer assignments)"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        product = get_object_or_404(Product, id=product_id)
        product.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def assign_product_to_customer_api(request):
    """Assign a tenant product to a customer (drag-and-drop)"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        customer_id = data.get('customer_id')
        if not product_id or not customer_id:
            return JsonResponse({'error': 'product_id and customer_id are required'}, status=400)

        product = get_object_or_404(Product, id=product_id)
        customer = get_object_or_404(Customer, id=customer_id)

        # Check if already assigned
        existing = CustomerProfile.objects.filter(customer=customer, product=product).first()
        if existing:
            return JsonResponse({'error': f'"{product.description}" is already assigned to this customer'}, status=400)

        profile = CustomerProfile.objects.create(
            tenant=tenant,
            customer=customer,
            product=product,
            description=product.description,
            unit_type=product.unit_type,
            pack_size=product.pack_size,
            sales_price=product.default_price,
            is_active=True,
        )
        # Include images from the tenant product
        images = []
        for img in product.images.all().order_by('slot'):
            images.append({'slot': img.slot, 'url': img.image.url})

        return JsonResponse({
            'success': True,
            'id': profile.id,
            'product_id': product.id,
            'description': profile.description,
            'unit_type': profile.unit_type,
            'pack_size': str(profile.pack_size) if profile.pack_size else '',
            'sales_price': str(profile.sales_price) if profile.sales_price else '0.00',
            'images': images,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def unassign_product_from_customer_api(request, profile_id):
    """Remove a product assignment from a customer"""
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        profile = get_object_or_404(CustomerProfile, id=profile_id)
        profile.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# VIEW ORDERS
# =============================================================================

@login_required
def profile_orders_list(request):
    from ..models import TenantUser
    tenant = get_current_tenant()
    tenant_users = TenantUser.objects.filter(tenant=tenant).select_related('user').order_by('user__username')
    users = [{'id': tu.user.id, 'name': tu.user.username} for tu in tenant_users]
    return render(request, 'core/Orders/profile_orders_list.html', {'users': users})


@login_required
def get_profile_orders_api(request):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)

    completed = request.GET.get('completed', 'false') == 'true'
    orders = SalesOrder.objects.filter(is_completed=completed).select_related(
        'customer', 'assigned_to', 'completed_by'
    ).prefetch_related('items').order_by('-id')

    data = []
    for o in orders:
        total = sum(float(item.amount or 0) for item in o.items.all())
        data.append({
            'soid': o.id,
            'order_number': o.order_number,
            'customer': o.customer_name or '',
            'dispatch_date': o.order_date.strftime('%m/%d/%Y') if o.order_date else '',
            'total': total,
            'customerpo': o.po_number or '',
            'comments': o.notes or '',
            'assigned_to_id': o.assigned_to_id,
            'assigned_to_name': o.assigned_to.username if o.assigned_to else '',
            'completed_at': o.completed_at.strftime('%m/%d/%Y %I:%M %p') if o.completed_at else '',
            'completed_by': o.completed_by.username if o.completed_by else '',
        })

    return JsonResponse({'orders': data})


@login_required
def get_profile_order_items_api(request, soid):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)

    so = get_object_or_404(SalesOrder, id=soid)
    items = SalesOrderItem.objects.filter(sales_order=so).order_by('sort_order', 'id')
    data = [{
        'description': i.description or '',
        'qty': float(i.quantity or 0),
        'pack': 1,
        'price': float(i.unit_price or 0),
        'total': float(i.amount or 0),
        'instructions': i.notes or '',
    } for i in items]

    return JsonResponse({'items': data})


@login_required
@require_POST
def assign_profile_order_api(request, soid):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        so = get_object_or_404(SalesOrder, id=soid)
        data = json.loads(request.body)
        user_id = data.get('user_id')

        if user_id:
            from django.contrib.auth.models import User as DjangoUser
            so.assigned_to = get_object_or_404(DjangoUser, id=user_id)
        else:
            so.assigned_to = None
        so.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def complete_profile_order_api(request, soid):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        from django.utils import timezone
        so = get_object_or_404(SalesOrder, id=soid)
        so.is_completed = True
        so.completed_at = timezone.now()
        so.completed_by = request.user
        so.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def uncomplete_profile_order_api(request, soid):
    tenant = get_current_tenant()
    if not tenant:
        return JsonResponse({'error': 'No tenant context'}, status=400)
    try:
        so = get_object_or_404(SalesOrder, id=soid)
        so.is_completed = False
        so.completed_at = None
        so.completed_by = None
        so.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
def public_profile_order_form(request, token):
    customer = get_object_or_404(Customer, public_token=token)
    profiles = CustomerProfile.objects.filter(customer=customer, is_active=True).order_by('description')
    return render(request, 'core/Orders/profile_order_form.html', {
        'customer': customer,
        'profiles': profiles,
    })
